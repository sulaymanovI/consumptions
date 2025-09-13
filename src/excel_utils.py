import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from config import CATEGORIES

def create_expenses_excel(expenses_data):
    """Создает Excel файл с расходами (оптимизированная версия)"""
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"
    
    # Заголовки столбцов
    headers = [
        "Пользователь",
        "Сумма (сум)",
        "Категория",
        "Описание", 
        "Комментарий",
        "Дата создания"
    ]
    
    # Упрощенные стили для оптимизации
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Заполняем данные (только основные поля для уменьшения размера)
    for row_num, expense in enumerate(expenses_data, 2):
        ws.cell(row=row_num, column=1, value=expense['first_name'])
        ws.cell(row=row_num, column=2, value=float(expense['amount']))
        ws.cell(row=row_num, column=3, value=CATEGORIES.get(expense['category'], expense['category']))
        ws.cell(row=row_num, column=4, value=expense['description'])
        ws.cell(row=row_num, column=5, value=expense['comment'] or "")
        
        # Форматируем дату
        if isinstance(expense['created_at'], datetime):
            date_str = expense['created_at'].strftime("%Y-%m-%d")
        else:
            date_str = str(expense['created_at'])[:10]  # Берем только дату
        ws.cell(row=row_num, column=6, value=date_str)
    
    # Настраиваем ширину столбцов
    column_widths = [15, 15, 20, 30, 30, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Создаем временный файл
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    # Проверяем размер файла
    file_size = os.path.getsize(temp_file.name)
    logging.info(f"Excel file size: {file_size} bytes")
    
    return temp_file.name

def cleanup_excel_file(file_path):
    """Удаляет временный Excel файл"""
    try:
        if os.path.exists(file_path):
            os.unlink(file_path)
    except:
        pass